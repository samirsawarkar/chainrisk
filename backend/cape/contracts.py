from typing import Optional

from openpyxl import load_workbook

from cape.sku_integrity import validate_config_skus_covered_in_scenario

SYSTEM_CONFIG_REQUIRED = ["nodes", "skus", "lead_times", "initial_inventory"]


def _int_field(value, field_label: str, errors: list[str]) -> Optional[int]:
    """Parse int for validation; append error and return None on failure."""
    try:
        return int(value)
    except (TypeError, ValueError):
        errors.append(f"{field_label} must be an integer")
        return None


def validate_system_config(config: dict) -> list[str]:
    errors: list[str] = []
    if not isinstance(config, dict):
        return ["system_config must be a JSON object"]

    for key in SYSTEM_CONFIG_REQUIRED:
        if key not in config:
            errors.append(f"Missing required field: {key}")
    if "nodes" in config and not isinstance(config["nodes"], list):
        errors.append("nodes must be an array")
    if "skus" in config and not isinstance(config["skus"], list):
        errors.append("skus must be an array")
    if "lead_times" in config and not isinstance(config["lead_times"], list):
        errors.append("lead_times must be an array")
    if "initial_inventory" in config and not isinstance(config["initial_inventory"], list):
        errors.append("initial_inventory must be an array")
    if errors:
        return errors

    node_ids = set()
    for i, n in enumerate(config.get("nodes", [])):
        if not isinstance(n, dict):
            errors.append(f"nodes[{i}] must be an object")
            continue
        nid = str(n.get("node_id", "")).strip()
        if nid:
            node_ids.add(nid)

    sku_ids = set()
    for i, s in enumerate(config.get("skus", [])):
        if not isinstance(s, dict):
            errors.append(f"skus[{i}] must be an object")
            continue
        sid = str(s.get("sku_id", "")).strip()
        if sid:
            sku_ids.add(sid)

    if not node_ids:
        errors.append("nodes must include at least one node_id")
    if not sku_ids:
        errors.append("skus must include at least one sku_id")
    if errors:
        return errors

    for i, node in enumerate(config.get("nodes", [])):
        if not isinstance(node, dict):
            continue
        nid = str(node.get("node_id", "UNKNOWN")).strip() or "UNKNOWN"
        cap = _int_field(node.get("capacity_units", 0) or 0, f"nodes[{i}].capacity_units", errors)
        if cap is not None and cap <= 0:
            errors.append(f"node {nid} has non-positive capacity_units")

    for i, arc in enumerate(config.get("lead_times", [])):
        if not isinstance(arc, dict):
            errors.append(f"lead_times[{i}] must be an object")
            continue
        frm = str(arc.get("from_node", "")).strip()
        to = str(arc.get("to_node", "")).strip()
        if frm not in node_ids or to not in node_ids:
            errors.append(f"lead_time arc {frm}->{to} references unknown node")
        lt = _int_field(arc.get("lead_time_ticks", 0) or 0, f"lead_times[{i}].lead_time_ticks", errors)
        if lt is not None and lt < 0:
            errors.append(f"lead_time arc {frm}->{to} has negative lead_time_ticks")

    for i, row in enumerate(config.get("initial_inventory", [])):
        if not isinstance(row, dict):
            errors.append(f"initial_inventory[{i}] must be an object")
            continue
        node_id = str(row.get("node_id", "")).strip()
        sku_id = str(row.get("sku_id", "")).strip()
        if node_id not in node_ids:
            errors.append(f"initial_inventory references unknown node_id: {node_id}")
        if sku_id not in sku_ids:
            errors.append(f"initial_inventory references unknown sku_id: {sku_id}")
        oh = _int_field(row.get("on_hand", 0) or 0, f"initial_inventory[{i}].on_hand", errors)
        if oh is not None and oh < 0:
            errors.append(f"initial_inventory has negative on_hand for {node_id}:{sku_id}")
    return errors


def normalize_scenario_rows(rows: list[dict]) -> list[dict]:
    normalized = []
    for row in rows:
        sku_value = row.get("sku", row.get("sku_id", ""))
        demand_value = row.get("demand", row.get("quantity", 0))
        normalized.append(
            {
                "time": int(row.get("time", 0)),
                "sku": str(sku_value).strip(),
                "demand": int(demand_value),
                "shock": int(row.get("shock", 0)),
            }
        )
    return normalized


def parse_json_scenario(payload: dict) -> list[dict]:
    if "demand" in payload and isinstance(payload["demand"], list):
        return normalize_scenario_rows(payload["demand"])
    if "scenario_events" in payload and isinstance(payload["scenario_events"], list):
        return normalize_scenario_rows(payload["scenario_events"])
    raise ValueError("JSON scenario must include 'demand' or 'scenario_events' array")


def parse_excel_to_json(uploaded_file) -> list[dict]:
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    header = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in ws[1]]
    required = ["time", "sku", "demand", "shock"]
    missing = [col for col in required if col not in header]
    if missing:
        raise ValueError(f"Missing scenario columns: {', '.join(missing)}")

    idx = {name: header.index(name) for name in required}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v in (None, "") for v in row):
            continue
        rows.append(
            {
                "time": row[idx["time"]],
                "sku": row[idx["sku"]],
                "demand": row[idx["demand"]],
                "shock": row[idx["shock"]],
            }
        )
    return normalize_scenario_rows(rows)


def compute_decision_signal(metrics: dict) -> dict:
    capacity_peak = 0.0
    backlog = int(metrics.get("system_backlog", 0))
    amp_ratios = metrics.get("amplification_ratios") or {}
    explain = metrics.get("explainability") or {}
    dist_over_ret = float(amp_ratios.get("dist_over_ret", 1.0))
    mfg_over_dist = float(amp_ratios.get("mfg_over_dist", 1.0))
    sup_over_mfg = float(amp_ratios.get("sup_over_mfg", 1.0))
    amp_peak = max(dist_over_ret, mfg_over_dist, sup_over_mfg)
    amplification_location = "none"
    causing_node = "none"
    if amp_peak == dist_over_ret and amp_peak > 1.0:
        amplification_location = "DIST/RET"
        causing_node = "DIST-01"
    elif amp_peak == mfg_over_dist and amp_peak > 1.0:
        amplification_location = "MFG/DIST"
        causing_node = "MFG-01"
    elif amp_peak == sup_over_mfg and amp_peak > 1.0:
        amplification_location = "SUP/MFG"
        causing_node = "SUP-01"
    if metrics.get("capacity_utilization"):
        capacity_peak = max(float(v) for v in metrics["capacity_utilization"].values())
    instability_index = float(metrics.get("instability_index", 1.0))
    bullwhip_detected = instability_index > 2.0 or amp_peak > 1.5
    status = "stable"
    if backlog > 0 or bullwhip_detected:
        status = "warning"
    if bullwhip_detected or capacity_peak > 90:
        status = "critical"
    root_cause = "baseline_variation"
    if bullwhip_detected:
        root_cause = "demand amplification + delay"
    elif capacity_peak > 90:
        root_cause = "capacity_saturation"
    elif backlog > 0:
        root_cause = "unfilled_demand_backlog"
    recommendation = "maintain_current_allocation"
    if "demand amplification" in root_cause or root_cause == "demand_amplification":
        recommendation = "reduce_high_variance_sku_allocation_by_12_percent"
    elif root_cause == "capacity_saturation":
        recommendation = "rebalance_capacity_or_shift_orders_to_lower_utilized_nodes"
    elif root_cause == "unfilled_demand_backlog":
        recommendation = "rebalance_allocation_toward_backlogged_skus_and_nodes"
    spike_sku = explain.get("spike_sku") or "UNKNOWN_SKU"
    spike_from = int(explain.get("spike_from") or 0)
    spike_to = int(explain.get("spike_to") or 0)
    peak_capacity_node = explain.get("peak_capacity_node") or causing_node
    peak_capacity_util = float(explain.get("peak_capacity_utilization") or capacity_peak)
    low_allocation_sku = explain.get("low_allocation_sku") or spike_sku

    if bullwhip_detected:
        root_cause = f"{spike_sku} demand spike caused {round(peak_capacity_util, 1)}% capacity usage at {peak_capacity_node}"
        recommendation = f"Reduce {spike_sku} allocation by 10–15% to stabilize capacity at {peak_capacity_node}"

    evidence = [
        f"{spike_sku} demand increased from {spike_from} → {spike_to}",
        f"{peak_capacity_node} capacity utilization reached {round(peak_capacity_util, 1)}%",
        f"{low_allocation_sku} allocation dropped below stable target",
    ]

    return {
        "time": int(metrics["tick"]),
        "status": status,
        "bullwhip_detected": bullwhip_detected,
        "peak_node": causing_node if causing_node != "none" else "SUP-01",
        "instability_index": round(instability_index, 4),
        "capacity_utilization": round(capacity_peak, 2),
        "bullwhip_index": round(instability_index, 4),
        "amplification_ratios": {
            "dist_over_ret": round(dist_over_ret, 4),
            "mfg_over_dist": round(mfg_over_dist, 4),
            "sup_over_mfg": round(sup_over_mfg, 4),
        },
        "amplification_location": amplification_location,
        "causing_node": causing_node,
        "root_cause": root_cause,
        "evidence": evidence,
        "impact": f"₹{abs(int(metrics['net_margin_impact'])):,} margin erosion",
        "recommendation": recommendation,
    }


def check_input_consistency(system_config: dict, scenario_events: list[dict]) -> dict:
    errors = []
    warnings = []
    suggestions = []

    cfg_errors = validate_system_config(system_config)
    errors.extend(cfg_errors)
    if cfg_errors:
        return {"valid": False, "errors": errors, "warnings": warnings, "suggestions": suggestions}

    config_skus = {str(item.get("sku_id", "")).strip() for item in system_config.get("skus", []) if item.get("sku_id")}
    scenario_skus = {str(item.get("sku", "")).strip() for item in scenario_events if item.get("sku")}
    missing_skus = sorted(sku for sku in scenario_skus if sku not in config_skus)
    if missing_skus:
        errors.append(f"Scenario contains SKUs not in system config: {', '.join(missing_skus)}")

    errors.extend(validate_config_skus_covered_in_scenario(system_config, scenario_events))

    node_ids = {str(node.get("node_id", "")).strip() for node in system_config.get("nodes", []) if node.get("node_id")}
    required_chain = {"SUP-01", "MFG-01", "DIST-01", "RET-01"}
    if not required_chain.issubset(node_ids):
        warnings.append("Recommended baseline chain missing one or more nodes: SUP-01, MFG-01, DIST-01, RET-01")
        suggestions.append("Add full 4-stage chain nodes for realistic bullwhip simulation.")

    if len(scenario_events) == 0:
        errors.append("Scenario events are empty.")
    else:
        max_time = max(int(row.get("time", 0)) for row in scenario_events)
        if max_time < 3:
            warnings.append("Scenario horizon is very short; instability effects may not appear.")
            suggestions.append("Use at least 8-12 time steps for meaningful behavior.")

    return {"valid": len(errors) == 0, "errors": errors, "warnings": warnings, "suggestions": suggestions}


def build_assistant_response(
    question: str,
    system_config: dict,
    scenario_events: list[dict],
    decision_signal: Optional[dict] = None,
) -> dict:
    q = (question or "").lower()
    consistency = check_input_consistency(system_config or {}, scenario_events or [])
    if not consistency["valid"]:
        return {
            "summary": "Your setup has blocking issues.",
            "recommendation": "Fix input consistency errors before running.",
            "details": consistency,
        }

    if "sku" in q:
        scenario_skus = sorted({row.get("sku") for row in scenario_events})
        return {
            "summary": f"Scenario includes SKUs: {', '.join(scenario_skus)}.",
            "recommendation": "Prioritize high-margin SKUs during shock windows and watch capacity saturation.",
            "details": {"sku_count": len(scenario_skus)},
        }

    if "capacity" in q or "bottleneck" in q:
        return {
            "summary": "Capacity bottlenecks are likely during demand spikes.",
            "recommendation": "Reduce allocations for volatile SKUs by 10-15% when utilization crosses 90%.",
            "details": {"rule": "utilization>90 => rebalance"},
        }

    if q.strip() in {"hi", "hey", "hello"}:
        status = (decision_signal or {}).get("status", "unknown")
        return {
            "summary": f"System status is {status}. Ask about a node, SKU, or tick range for specifics.",
            "recommendation": "Try: 'What happened at T=4-6 for SKU B?' or 'What did MFG-01 decide at T=5?'",
            "details": {"hints": ["node_id", "sku", "tick_range"]},
        }

    if "tell all" in q or "everything" in q or "full" in q:
        scenario_rows = len(scenario_events or [])
        sku_count = len(system_config.get("skus", []) if isinstance(system_config, dict) else [])
        node_count = len(system_config.get("nodes", []) if isinstance(system_config, dict) else [])
        status = (decision_signal or {}).get("status", "unknown")
        root_cause = (decision_signal or {}).get("root_cause", "unknown")
        recommendation = (decision_signal or {}).get("recommendation", "review report")
        return {
            "summary": f"Full snapshot: {status} state with root cause '{root_cause}'.",
            "recommendation": recommendation,
            "details": {
                "nodes": node_count,
                "skus": sku_count,
                "scenario_rows": scenario_rows,
                "next_questions": [
                    "Where did amplification happen?",
                    "Which node created backlog first?",
                    "What changed between T=4 and T=6?",
                ],
            },
        }

    if decision_signal:
        return {
            "summary": (
                f"Current status is {decision_signal.get('status', 'unknown')} "
                f"(cause: {decision_signal.get('root_cause', 'unknown')})."
            ),
            "recommendation": decision_signal.get("recommendation", "Review decision center signals."),
            "details": {
                "impact": decision_signal.get("impact"),
                "causing_node": decision_signal.get("causing_node"),
                "amplification_location": decision_signal.get("amplification_location"),
            },
        }

    return {
        "summary": "Input looks valid and ready for simulation.",
        "recommendation": "Run simulation, then inspect Decision Center for impact and root cause.",
        "details": {"scenario_rows": len(scenario_events)},
    }
